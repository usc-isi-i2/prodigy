#!/usr/bin/env python3
"""Consolidate every topology_feature_ssl data source into one workbook.

Raw_* sheets = verbatim source CSVs (the shared benchmark CSVs filtered to this
experiment's model family; the tfssl-only diagnostic CSVs in full). Every summary /
pivot sheet holds values computed in Python from those same Raw_* rows (test split),
== the AVERAGEIFS you would write over each Raw_* sheet, re-derivable via
analyze_matched40k.py. Values not live formulas because LibreOffice is unavailable in
this environment to validate cross-sheet formulas; the Raw_* sheets keep it auditable.

Reproduce:  python scripts/experiments/topology_feature_ssl/build_results_xlsx.py
"""
from __future__ import annotations
import csv
import statistics
from pathlib import Path
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
REPO = HERE.parents[2]
PLOT = REPO / "scripts/plotting"
TF = PLOT / "topology_feature_ssl/data"
OUT = HERE / "topology_feature_ssl_results.xlsx"

ARMS = ["B0_40k", "B1_40k", "E1_40k", "E2_40k", "E2b_40k", "E4_40k", "E4r_40k"]
ARM_DISP = {a: a[:-4] for a in ARMS}  # B0_40k -> B0
ROLE = {
    "B0_40k": "control — NM, mean-agg encoder, no aug",
    "B1_40k": "augmentation lever — NM + 30% random-feature corruption",
    "E1_40k": "encoder inputs — + directed in/out/log-degree",
    "E2_40k": "encoder aggregation — sum/PNA, directed split, multi-readout",
    "E2b_40k": "mechanism probe — E2 minus conv BatchNorm",
    "E4_40k": "objective — multi-task MFR⊕dirLP⊕structural (simultaneous)",
    "E4r_40k": "objective — multi-task, per-episode rotation",
}
REG6 = ["followers_count", "friends_count", "statuses_count",
        "favourites_count", "listed_count", "account_age_days"]
CLS_DS = ["election2020", "twibot20"]
GRAPH_DS = ["covid19_twitter", "midterm", "twibot20", "ukr_rus_twitter"]
PROBE_RULES = ["count_threshold", "in_degree", "out_degree", "existence", "conjunction"]

# ---- model filter: keep only this experiment's family in the shared benchmark CSVs
FAMILY_EXTRA = {"task_transfer_covid_nm", "task_transfer_covid_fp",
                "task_transfer_covid_cl", "random_init"}


def in_family(model: str) -> bool:
    return model.split("_")[0] in {"B0", "B1", "E1", "E2", "E2b", "E4", "E4r"} or model in FAMILY_EXTRA


# ----------------------------------------------------------------------------- style
FONT = "Arial"
INK = "202124"
HDR_FILL = PatternFill("solid", fgColor="34405A")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color=INK)
SUB_FONT = Font(name=FONT, italic=True, size=10, color="5F6368")
BOLD = Font(name=FONT, bold=True, color=INK, size=11)
BASE = Font(name=FONT, color=INK, size=11)
WIN_FILL = PatternFill("solid", fgColor="E4F0E4")   # arm that wins its task
FLOOR_FILL = PatternFill("solid", fgColor="F5F6F8")  # baseline/floor rows
MEAN_FILL = PatternFill("solid", fgColor="EEF0FF")
thin = Side(style="thin", color="D0D3D8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
NUM = "0.000"
DELTA = "+0.000;\\-0.000;0.000"
CEN = Alignment(horizontal="center", vertical="center")
LEF = Alignment(horizontal="left", vertical="center")
TOP = Alignment(vertical="top", wrap_text=True)


def read_csv(p):
    with open(p) as fh:
        return list(csv.reader(fh))


def num(v):
    try:
        return int(v) if str(v).strip().lstrip("-").isdigit() else float(v)
    except (ValueError, AttributeError):
        return v


# ----------------------------------------------------------------------------- raw load
def load_rows(path, model_filter=False):
    rows = read_csv(path)
    if model_filter:
        rows = [rows[0]] + [r for r in rows[1:] if r and in_family(r[0])]
    return rows


REG = load_rows(PLOT / "node_regression/data/node_regression.csv", model_filter=True)
CLS = load_rows(PLOT / "node_classification/data/node_classification.csv", model_filter=True)
SLP = load_rows(PLOT / "static_link_prediction/data/static_link_prediction.csv", model_filter=True)
PROBES = read_csv(TF / "capability_probes_40k.csv")
ABL = read_csv(TF / "ablation_2x2_40k.csv")
BUD = read_csv(TF / "budget_sweep.csv")
LEAK = read_csv(TF / "leakage_baseline.csv")
FLOOR = read_csv(PLOT / "node_regression/data/features_only_floor.csv")
TRIV = read_csv(TF / "trivial_baselines_small.csv")
PROBES_E = read_csv(TF / "capability_probes.csv")
ABL_E = read_csv(TF / "ablation_2x2.csv")

wb = Workbook()

# raw-sheet registry: sheetname -> (rows, numeric_col_indices_0based)
RAW_SPECS = {
    "Raw_regression": (REG, {2, 7, 8, 9, 10, 11}),
    "Raw_classification": (CLS, {2, 7, 8, 9}),
    "Raw_static_link_prediction": (SLP, {2, 7, 8, 9}),
    "Raw_capability_probes_40k": (PROBES, {2}),
    "Raw_ablation_2x2_40k": (ABL, {5, 6, 7}),
    "Raw_budget_sweep": (BUD, {1, 2}),
    "Raw_leakage_baseline": (LEAK, {2, 3, 4, 5}),
    "Raw_features_floor": (FLOOR, {2, 3, 4, 5}),
    "Raw_trivial_baselines": (TRIV, {4, 5}),
    "Raw_capability_probes_early": (PROBES_E, {2}),
    "Raw_ablation_2x2_early": (ABL_E, {5, 6, 7}),
}
# In-memory index for computing summary values. Values (not live formulas) because
# LibreOffice is unavailable in this environment to validate cross-sheet formulas; each
# summary cell = mean over the named Raw_* sheet (== AVERAGEIFS), auditable there.
RAWIDX = {name: (rows[0], rows[1:]) for name, (rows, _) in RAW_SPECS.items()}


def agg(sheet, value_h, **crit):
    """Mean of value_h over rows of `sheet` matching every criterion (== AVERAGEIFS).
    Returns None when nothing matches."""
    header, rows = RAWIDX[sheet]
    vi = header.index(value_h)
    ci = {k: header.index(k) for k in crit}
    xs = []
    for r in rows:
        if all(str(r[ci[k]]) == str(v) for k, v in crit.items()):
            try:
                xs.append(float(r[vi]))
            except (ValueError, IndexError):
                pass
    return statistics.fmean(xs) if xs else None


aifs = agg  # call sites read as "the AVERAGEIFS value"


# ----------------------------------------------------------------------------- helpers
def title(ws, txt, sub=None):
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, txt).font = TITLE_FONT
    if sub:
        ws.cell(2, 1, sub).font = SUB_FONT


def style_hdr(ws, row, cols, start=1):
    for k, h in enumerate(cols, start=start):
        c = ws.cell(row, k, h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CEN; c.border = BORDER


def put(ws, r, c, v, *, fmt=None, font=BASE, fill=None, align=CEN, border=True, comment=None):
    cell = ws.cell(r, c, v)
    cell.font = font; cell.alignment = align
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if border:
        cell.border = BORDER
    if comment:
        cell.comment = Comment(comment, "results")
    return cell


def autosize(ws, rows, cap=48):
    for j in range(1, len(rows[0]) + 1):
        w = max([len(str(rows[0][j - 1]))] + [len(str(r[j - 1])) for r in rows[1:] if j - 1 < len(r)] or [0])
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 8), cap)


# ============================================================== Python ground truth
def pmean(rows, value_h, **crit):
    h = rows[0]; idx = {k: h.index(k) for k in h}
    xs = []
    for r in rows[1:]:
        if all(r[idx[k]] == v for k, v in crit.items()):
            try:
                xs.append(float(r[idx[value_h]]))
            except ValueError:
                pass
    return statistics.fmean(xs) if xs else None


reg_m = {a: pmean(REG, "spearman", model=a, split="test") for a in ARMS}
cls_m = {a: pmean(CLS, "roc_auc", model=a, split="test") for a in ARMS}
slp_m = {a: pmean(SLP, "roc_auc", model=a, split="test") for a in ARMS}

# ============================================================== About
ws = wb.active
ws.title = "About"
ws.sheet_view.showGridLines = False
lines = [
    ("topology_feature_ssl — consolidated results", TITLE_FONT),
    ("Can one frozen SSL pretext transfer to BOTH feature tasks (cls, regression) and a topological task (static-LP)?", SUB_FONT),
    ("Answer: NO arm clears both. Encoder axis — E1 wins features, E2 wins topology (different arms). Objective axis — E4 multi-task fails both. Joint goal unreached across all three levers.", SUB_FONT),
    ("", None),
    ("Arms (matched 40k-episode budget, true state_dict_40000, single seed)", "HEAD"),
    ("B0", "control — Neighborhood-Matching (NM), mean-agg GraphSAGE, bio features only, undirected, no augmentation. The reference."),
    ("B1", "AUGMENTATION lever — B0 + NR0.3 (30% of nodes get a random real feature vector). Hyp: corrupting the feature shortcut forces topology use without changing the objective."),
    ("E1", "ENCODER-INPUT lever — B0 + directed in/out/log-degree node inputs (directed3). Hyp: makes degree/direction representable. Leakage caveat: followers≈in-deg, statuses≈out-deg → must beat raw_degree."),
    ("E2", "ENCODER-AGGREGATION lever (composite) — mean→sum/PNA, in/out neighbors aggregated separately, readout mean→mean⊕sum⊕max. Hyp: makes counts/existence/conjunctions representable via aggregation."),
    ("E2b", "MECHANISM probe — E2 minus conv BatchNorm (no_bn_encoder). Hyp: BN washes out sum-aggregation's count magnitude; removing it should lift count/degree probes."),
    ("E4 / E4r", "OBJECTIVE lever — multi-task MFR⊕directed-LP⊕structural on E2's encoder (E4=simultaneous, E4r=rotation). RUN 2026-07-13. Both FAIL: regression negative, classification crashed below the floor, static-LP ≤ E2. E3 (masked-feature-only) folded in as E4's MFR head."),
    ("", None),
    ("Evaluation (frozen encoder)", "HEAD"),
    ("Classification", "ROC-AUC, 10-shot. Datasets: election2020, twibot20 (both held-out → pure transfer)."),
    ("Regression", "Spearman ρ, 10-shot, 6-target profile panel. Datasets: covid19, midterm, ukr_rus (in-domain) + twibot20 (held-out)."),
    ("Static link prediction", "ROC-AUC, 0-shot. Datasets: covid19, midterm, ukr_rus (in-domain) + twibot20 (held-out). The direct topological task."),
    ("Diagnostics", "Capability probes (linear-probe AUC on planted single rules, chance 0.50) + 2×2 feature/edge ablation."),
    ("", None),
    ("Trivial floors — an arm 'improves performance' only if it beats these", "HEAD"),
    ("raw_feat", "linear probe of raw bio features onto each target, NO encoder (Raw_features_floor for regression; Raw_trivial_baselines for classification)."),
    ("raw_degree", "linear probe of [in_deg, out_deg, log_deg] onto each target — the leakage control for E1/E2 (Raw_leakage_baseline). 3 targets only."),
    ("", None),
    ("Pretrain corpus", "HEAD"),
    ("Merged graph", "3-way merged retweet graph ukr_rus + covid19 + midterm (~34M nodes). One merged graph removes the pretrain-dataset multiplier. Within-source, source-balanced episode sampling."),
    ("Budget = 40k", "NM anti-scales on regression: E1 peaks at 40k then degrades to 110k (see Budget_sweep). All arms compared at a true state_dict_40000 checkpoint (E2/E2b use epochs:5 for the trainer off-by-one)."),
    ("", None),
    ("This workbook", "HEAD"),
    ("Summary/pivot sheets", "Summary_matched40k, Regression, Regression_detail, Classification, Static_link_prediction, Capability_probes, Ablation_2x2, Budget_sweep, Free_preview."),
    ("Summary values", "Each summary cell = mean over the named Raw_* sheet (test split), i.e. AVERAGEIFS computed in Python (LibreOffice unavailable here to validate live formulas). Re-derivable via analyze_matched40k.py. Green fill = arm that wins that task; grey = trivial floor."),
    ("Raw_* sheets", "Verbatim source CSVs. The 3 shared benchmark CSVs (regression/classification/static-LP) are FILTERED to this experiment's model family; other experiments' models were dropped. Diagnostic CSVs are full."),
    ("Model family kept", "B0*/B1*/E1*/E2*/E2b* (incl. _40k matched-eval + _step* budget-sweep checkpoints), free-preview proxies task_transfer_covid_{nm,fp,cl}, and random_init. Excluded: muc10k and all non-tfssl models."),
    ("Reproduce", "python scripts/experiments/topology_feature_ssl/build_results_xlsx.py  (then recalc.py to cache formula values)"),
    ("Built", "2026-07-12"),
    ("", None),
    ("Deviations / caveats (stated plainly)", "HEAD"),
    ("Objective axis run — fails", "E4 (simultaneous) + E4r (rotation) executed 2026-07-13; both fail the joint bar (reg negative, cls crashed, static-LP ≤ E2). Even after clipping the structural target, the struct term dominated the loss (weighted struct 1.5 ≫ lp 0.13 ≫ mfr 0.02), so the encoder over-fit degree and lost feature content. E4 reg is over the 3 swept targets only (followers/statuses/account_age)."),
    ("2×2 uninformative", "Feature-task metrics sit near zero for these arms, so the retained-fraction (metric ÷ intact) explodes/inverts. Reported for completeness; the topological read rests on Static_link_prediction + Capability_probes."),
    ("Baseline provenance", "The full-panel raw_feat/raw_degree rendered on the compute node (trivial_baselines.csv, leakage_baseline_6panel.csv) were not synced locally; the local floors here give an identical qualitative verdict."),
    ("Single seed", "Frozen-probe eval episodes are seeded per-split, not by --seed. Lean on cross-dataset agreement, not single-dataset deltas."),
]
r = 1
for a, b in lines:
    if b is None:
        r += 1; continue
    if isinstance(b, Font):
        ws.cell(r, 1, a).font = b
    elif b == "HEAD":
        ws.cell(r, 1, a).font = BOLD
    else:
        c = ws.cell(r, 1, a); c.font = BOLD; c.alignment = Alignment(vertical="top")
        c2 = ws.cell(r, 2, b); c2.font = BASE; c2.alignment = TOP
    r += 1
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 118

# ============================================================== Summary_matched40k
ws = wb.create_sheet("Summary_matched40k")
title(ws, "Matched-40k transfer — mean over datasets (test split)",
      "Feature axis = cls AUC / regression ρ; topological axis = static-LP AUC. Joint bar = min(cls, sLP), chance 0.50.")
style_hdr(ws, 4, ["Arm", "Classification AUC", "Regression ρ", "Static-LP AUC", "Joint min(cls,sLP)", "Role"])
ws.cell(4, 3).comment = Comment("Mean over 6 targets × 4 datasets (Raw_regression, split=test).", "results")
ws.cell(4, 5).comment = Comment("min(feature=cls AUC, topological=sLP AUC). Regression excluded from the joint bar (different scale).", "results")
for i, arm in enumerate(ARMS):
    rr = 5 + i
    put(ws, rr, 1, ARM_DISP[arm], font=BOLD, align=LEF)
    put(ws, rr, 2, aifs("Raw_classification", "roc_auc", model=arm, split="test"), fmt=NUM)
    put(ws, rr, 3, aifs("Raw_regression", "spearman", model=arm, split="test"), fmt=NUM)
    put(ws, rr, 4, aifs("Raw_static_link_prediction", "roc_auc", model=arm, split="test"), fmt=NUM)
    put(ws, rr, 5, min(cls_m[arm], slp_m[arm]), fmt=NUM)
    put(ws, rr, 6, ROLE[arm], font=BASE, align=LEF)
# highlight task winners (green): reg -> E1, slp -> E2
put(ws, 7, 3, aifs("Raw_regression", "spearman", model="E1_40k", split="test"), fmt=NUM, font=BOLD, fill=WIN_FILL)  # E1 reg
put(ws, 8, 4, aifs("Raw_static_link_prediction", "roc_auc", model="E2_40k", split="test"), fmt=NUM, font=BOLD, fill=WIN_FILL)  # E2 slp
for k, t in enumerate([
    "No arm clears the joint bar. Encoder axis: E1 wins features (regression), E2 wins topology (static-LP 0.76) — different arms. Objective axis: E4/E4r multi-task FAIL both.",
    "E1 is the only arm with usable regression; E2 the best static-LP (0.76). B1, E2b, E4r fall below chance on static-LP.",
    "E4's joint min(cls,slp)=0.45 and E4r's=0.23 are WORSE than every NM arm — the multi-task objective degraded the encoder (struct term dominated; MFR too weak to hold features).",
]):
    ws.cell(13 + k, 1, "• " + t).font = BASE
for col, w in zip("ABCDEF", [8, 18, 14, 14, 18, 58]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ============================================================== Regression
ws = wb.create_sheet("Regression")
title(ws, "Node regression — Spearman ρ (test, mean over 4 datasets, 10-shot)",
      "Floors: raw_feat (bio features, no encoder) and raw_degree (leakage). E1 is the only arm positive on all 6 targets.")
style_hdr(ws, 4, ["Arm / floor"] + [t.replace("_count", "").replace("_days", "") for t in REG6])
# floor rows first
put(ws, 5, 1, "raw_degree (leak)", font=BOLD, align=LEF, fill=FLOOR_FILL)
for j, t in enumerate(REG6):
    v = aifs("Raw_leakage_baseline", "spearman", target=t) if t in {"followers_count", "statuses_count", "account_age_days"} else None
    put(ws, 5, 2 + j, v if v is not None else "—", fmt=NUM, fill=FLOOR_FILL, font=BASE)
put(ws, 6, 1, "raw_feat (floor)", font=BOLD, align=LEF, fill=FLOOR_FILL)
for j, t in enumerate(REG6):
    put(ws, 6, 2 + j, aifs("Raw_features_floor", "spearman", target=t), fmt=NUM, fill=FLOOR_FILL, font=BASE)
for i, arm in enumerate(ARMS):
    rr = 7 + i
    put(ws, rr, 1, ARM_DISP[arm], font=BOLD, align=LEF)
    for j, t in enumerate(REG6):
        v = aifs("Raw_regression", "spearman", model=arm, target=t, split="test")
        put(ws, rr, 2 + j, v if v is not None else "—",
            fmt=NUM, fill=WIN_FILL if arm == "E1_40k" else None, font=BOLD if arm == "E1_40k" else BASE)
put(ws, 15, 1, "E1 (only positive arm) beats raw_feat on followers/friends/statuses/account_age; leakage-free win: account_age (0.12 vs 0.02). "
    "E4/E4r are NEGATIVE on all 3 swept targets — the multi-task objective loses feature content.", font=SUB_FONT, align=LEF, border=False)
put(ws, 16, 1, "“—” = not measured. E2b/E4/E4r ran only the 3 structure/age targets (followers, statuses, account_age).",
    font=SUB_FONT, align=LEF, border=False)
ws.column_dimensions["A"].width = 18
for col in "BCDEFG":
    ws.column_dimensions[col].width = 11
ws.freeze_panes = "B5"

# ============================================================== Regression_detail
ws = wb.create_sheet("Regression_detail")
title(ws, "Regression per dataset — the 3 structure-linked targets",
      "Shows where E1's advantage is genuine (account_age, content-linked) vs degree-explained (statuses).")
row = 4
for tgt in ["followers_count", "statuses_count", "account_age_days"]:
    ws.cell(row, 1, tgt).font = BOLD
    style_hdr(ws, row + 1, ["Arm"] + GRAPH_DS)
    for i, arm in enumerate(ARMS):
        rr = row + 2 + i
        put(ws, rr, 1, ARM_DISP[arm], font=BOLD, align=LEF)
        for j, ds in enumerate(GRAPH_DS):
            put(ws, rr, 2 + j, aifs("Raw_regression", "spearman", model=arm, dataset=ds, target=tgt, split="test"),
                fmt=NUM, fill=WIN_FILL if arm == "E1_40k" else None, font=BOLD if arm == "E1_40k" else BASE)
    row += 2 + len(ARMS) + 1
ws.column_dimensions["A"].width = 14
for col in "BCDE":
    ws.column_dimensions[col].width = 16
ws.freeze_panes = "B4"

# ============================================================== Classification
ws = wb.create_sheet("Classification")
title(ws, "Node classification — ROC-AUC (test, 10-shot, held-out datasets)",
      "NM arms (B0–E2b) cluster ~0.78; the multi-task E4/E4r CRASH below the raw_feat floor (E4 0.45, twibot20 0.38 < chance).")
style_hdr(ws, 4, ["Arm / floor"] + CLS_DS)
put(ws, 5, 1, "raw_feat (floor)", font=BOLD, align=LEF, fill=FLOOR_FILL)
for j, ds in enumerate(CLS_DS):
    put(ws, 5, 2 + j, aifs("Raw_trivial_baselines", "metric", task="classification", dataset=ds), fmt=NUM, fill=FLOOR_FILL)
for i, arm in enumerate(ARMS):
    rr = 6 + i
    put(ws, rr, 1, ARM_DISP[arm], font=BOLD, align=LEF)
    for j, ds in enumerate(CLS_DS):
        put(ws, rr, 2 + j, aifs("Raw_classification", "roc_auc", model=arm, dataset=ds, split="test"), fmt=NUM)
ws.column_dimensions["A"].width = 18
for col in "BC":
    ws.column_dimensions[col].width = 15
ws.freeze_panes = "B5"

# ============================================================== Static_link_prediction
ws = wb.create_sheet("Static_link_prediction")
title(ws, "Static link prediction — ROC-AUC (test, 0-shot)  —  the topological task",
      "E2 (count-aware aggregation) best (0.76). E4 middling (0.66, ≤ E2 and B0). B1, E2b, E4r fall BELOW chance (0.50).")
style_hdr(ws, 4, ["Arm"] + GRAPH_DS + ["MEAN"])
for i, arm in enumerate(ARMS):
    rr = 5 + i
    win = arm == "E2_40k"
    put(ws, rr, 1, ARM_DISP[arm], font=BOLD, align=LEF, fill=WIN_FILL if win else None)
    for j, ds in enumerate(GRAPH_DS):
        put(ws, rr, 2 + j, aifs("Raw_static_link_prediction", "roc_auc", model=arm, dataset=ds, split="test"),
            fmt=NUM, fill=WIN_FILL if win else None, font=BOLD if win else BASE)
    put(ws, rr, 6, slp_m[arm], fmt=NUM, font=BOLD, fill=WIN_FILL if win else MEAN_FILL)
put(ws, 13, 1, "Chance = 0.50. E2 mean 0.76 ≫ B0 0.68 ≈ E1 0.66 ≈ E4 0.66. Aggregating over neighbours (E2) helps LP; "
    "the E4 multi-task objective does NOT beat clean NM here, and E4r collapses below chance.", font=SUB_FONT, align=LEF, border=False)
ws.column_dimensions["A"].width = 10
for col in "BCDEF":
    ws.column_dimensions[col].width = 15
ws.freeze_panes = "B5"

# ============================================================== Capability_probes
ws = wb.create_sheet("Capability_probes")
title(ws, "Capability probes — linear-probe AUC (chance 0.50)",
      "Can a planted single structural rule be linearly read from the frozen rep? B0/B1 ≈ chance everywhere.")
style_hdr(ws, 4, ["Arm"] + PROBE_RULES)
for i, arm in enumerate(ARMS):
    rr = 5 + i
    put(ws, rr, 1, ARM_DISP[arm], font=BOLD, align=LEF)
    for j, rule in enumerate(PROBE_RULES):
        put(ws, rr, 2 + j, aifs("Raw_capability_probes_40k", "roc_auc", arm=arm, rule=rule), fmt=NUM)
for k, t in enumerate([
    "E1 leads count/in-degree (largely passthrough of its degree INPUTS).",
    "E2 leads existence/conjunction — the multi-neighbour rules that require summation (a mean encoder cannot represent them).",
    "E2b lifts count/out-degree (BN-washout confirmed) — but that gain does NOT reach the tasks (see Static_link_prediction).",
    "E4/E4r probes sit BELOW chance (0.15–0.51) — the multi-task rep does not linearly encode the planted rules; consistent with its crashed downstream tasks.",
]):
    ws.cell(13 + k, 1, "• " + t).font = SUB_FONT
ws.column_dimensions["A"].width = 10
for col in "BCDEF":
    ws.column_dimensions[col].width = 16
ws.freeze_panes = "B5"

# ============================================================== Ablation_2x2
ws = wb.create_sheet("Ablation_2x2")
title(ws, "2×2 ablation — retained fraction (metric ÷ intact)  —  REPORTED AS UNINFORMATIVE",
      "Near-zero feature-task metrics make the ratio explode/invert. Kept for completeness; do not read as signal.")
row = 4
for task, lab in [("reg", "regression"), ("pl", "classification (pl)"), ("slp", "static-LP")]:
    ws.cell(row, 1, f"task = {task}  ({lab})").font = BOLD
    style_hdr(ws, row + 1, ["Arm", "random_feat", "rewired_edge", "both"])
    for i, arm in enumerate(ARMS):
        rr = row + 2 + i
        put(ws, rr, 1, ARM_DISP[arm], font=BOLD, align=LEF)
        for j, cond in enumerate(["random_feat", "rewired_edge", "both"]):
            v = aifs("Raw_ablation_2x2_40k", "retained", arm=arm, task=task, condition=cond)
            put(ws, rr, 2 + j, v if v is not None else "—", fmt=NUM)
    row += 2 + len(ARMS) + 1
put(ws, row, 1, "retained = ablated-metric ÷ intact-metric. Values ≫1 or <0 are artefacts of near-zero denominators, "
    "not effects. The topological read uses Static_link_prediction + Capability_probes instead.", font=SUB_FONT, align=LEF, border=False)
ws.column_dimensions["A"].width = 12
for col in "BCD":
    ws.column_dimensions[col].width = 15
ws.freeze_panes = "B4"

# ============================================================== Budget_sweep
ws = wb.create_sheet("Budget_sweep")
title(ws, "Budget sweep — why 40k (B0 / E1)",
      "NM anti-scales on regression: E1 peaks at 40k then degrades to 110k. Classification is flat across budget.")
steps = [20000, 40000, 60000, 110000]
row = 4
for task, lab in [("node_regression", "Regression ρ (mean over targets)"),
                  ("node_classification", "Classification AUC")]:
    ws.cell(row, 1, lab).font = BOLD
    style_hdr(ws, row + 1, ["Step"] + ["B0", "E1"])
    for i, st in enumerate(steps):
        rr = row + 2 + i
        put(ws, rr, 1, st, font=BOLD, align=LEF)
        for j, arm in enumerate(["B0", "E1"]):
            put(ws, rr, 2 + j, aifs("Raw_budget_sweep", "score", arm=arm, step=st, task=task),
                fmt=NUM, fill=WIN_FILL if (task == "node_regression" and st == 40000) else None)
    row += 2 + len(steps) + 1
put(ws, row, 1, "Regression peaks at 40k (E1 0.222) and erodes by 110k (0.142) — instance discrimination collapses "
    "continuous variation. Hence the matched-40k design.", font=SUB_FONT, align=LEF, border=False)
ws.column_dimensions["A"].width = 12
for col in "BC":
    ws.column_dimensions[col].width = 12
ws.freeze_panes = "B4"

# ============================================================== Free_preview
ws = wb.create_sheet("Free_preview")
title(ws, "Free preview — masked-feature (fp) vs NM, regression ρ (test)",
      "Zero-cost E3 pre-check on the existing covid checkpoints. fp does NOT beat NM → E3 not pre-validated.")
# build (dataset,target) list present for fp in the filtered regression rows
h = REG[0]; mi, di, ti, si = h.index("model"), h.index("dataset"), h.index("target"), h.index("split")
pairs = sorted({(r[di], r[ti]) for r in REG[1:] if r[mi] == "task_transfer_covid_fp" and r[si] == "test"})
style_hdr(ws, 4, ["dataset", "target", "nm", "fp", "fp − nm"])
fp_deltas = []
for i, (ds, tgt) in enumerate(pairs):
    rr = 5 + i
    nmv = aifs("Raw_regression", "spearman", model="task_transfer_covid_nm", dataset=ds, target=tgt, split="test")
    fpv = aifs("Raw_regression", "spearman", model="task_transfer_covid_fp", dataset=ds, target=tgt, split="test")
    d = (fpv - nmv) if (nmv is not None and fpv is not None) else None
    if d is not None:
        fp_deltas.append(d)
    put(ws, rr, 1, ds, font=BASE, align=LEF)
    put(ws, rr, 2, tgt, font=BASE, align=LEF)
    put(ws, rr, 3, nmv, fmt=NUM)
    put(ws, rr, 4, fpv, fmt=NUM)
    put(ws, rr, 5, d, fmt=DELTA)
mr = 5 + len(pairs)
put(ws, mr, 1, "mean(fp − nm)", font=BOLD, align=LEF, fill=MEAN_FILL)
put(ws, mr, 2, "", fill=MEAN_FILL)
put(ws, mr, 3, "", fill=MEAN_FILL); put(ws, mr, 4, "", fill=MEAN_FILL)
put(ws, mr, 5, statistics.fmean(fp_deltas), fmt=DELTA, font=BOLD, fill=MEAN_FILL,
    comment="Negative → masked-feature does not beat NM on regression. Matches the RESULTS.md free-preview verdict (−0.016).")
for col, w in zip("ABCDE", [18, 18, 11, 11, 12]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ============================================================== Raw sheets
def write_raw(name, rows, numeric):
    ws = wb.create_sheet(name)
    for j, hd in enumerate(rows[0], start=1):
        put(ws, 1, j, hd, font=HDR_FONT, fill=HDR_FILL)
    for i, rrow in enumerate(rows[1:], start=2):
        for j, v in enumerate(rrow, start=1):
            val = num(v) if (j - 1) in numeric else v
            c = put(ws, i, j, val, font=BASE, align=(CEN if (j - 1) in numeric else LEF), border=False)
            if isinstance(val, float):
                c.number_format = "0.0000####"
    ws.freeze_panes = "A2"
    autosize(ws, rows)


for name, (rows, numeric) in RAW_SPECS.items():
    write_raw(name, rows, numeric)

# ---- order
order = ["About", "Summary_matched40k", "Regression", "Regression_detail", "Classification",
         "Static_link_prediction", "Capability_probes", "Ablation_2x2", "Budget_sweep",
         "Free_preview"] + list(RAW_SPECS.keys())
wb._sheets.sort(key=lambda s: order.index(s.title))

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print("wrote", OUT)
print("sheets:", [s.title for s in wb._sheets])
print("\nWritten summary values (also auditable in the Raw_* sheets):")
print("  cls", {ARM_DISP[a]: round(cls_m[a], 3) for a in ARMS})
print("  reg", {ARM_DISP[a]: round(reg_m[a], 3) for a in ARMS})
print("  slp", {ARM_DISP[a]: round(slp_m[a], 3) for a in ARMS})
print("raw row counts:", {n: len(r) - 1 for n, (r, _) in RAW_SPECS.items()})
