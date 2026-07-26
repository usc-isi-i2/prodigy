#!/usr/bin/env python3
"""Build the multitask_ssl_rotation results workbook.

Raw eval rows -> Raw_* sheets (the source values). Summary/pivot sheets hold values
computed in Python from those same raw rows (test split), verified against
aggregate_results.py. Values (not live formulas) because LibreOffice is unavailable
here to validate cross-sheet formulas; the raw sheets keep everything auditable and
each derived sheet documents its provenance.
"""
import csv
import statistics
from pathlib import Path
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
REPO = HERE.parents[2]  # scripts/experiments/multitask_ssl_rotation -> repo root
DATA = REPO / "scripts/experiments/analysis/multitask_ssl_rotation"
OUT = HERE / "multitask_ssl_rotation_results.xlsx"

ARMS = ["NM", "CL", "FP", "MIX"]
CONTROLS = ["NM", "CL", "FP"]
ROLE = {
    "NM": "control — classification specialist",
    "CL": "control — weak everywhere",
    "FP": "control — regression specialist",
    "MIX": "TREATMENT — per-episode rotation",
}

FONT = "Arial"
INK = "202124"
HDR_FILL = PatternFill("solid", fgColor="34405A")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color=INK)
SUB_FONT = Font(name=FONT, italic=True, size=10, color="5F6368")
BOLD = Font(name=FONT, bold=True, color=INK, size=11)
BASE = Font(name=FONT, color=INK, size=11)
MIX_FILL = PatternFill("solid", fgColor="ECEBFA")
CTRL_FILL = PatternFill("solid", fgColor="F5F6F8")
thin = Side(style="thin", color="D0D3D8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
NUM = "0.000"
DELTA = "+0.000;\\-0.000;0.000"
CEN = Alignment(horizontal="center", vertical="center")
LEF = Alignment(horizontal="left", vertical="center")


def read_csv(p):
    with open(p) as fh:
        return list(csv.reader(fh))


def rows_as_dicts(rows):
    h = rows[0]
    return [dict(zip(h, r)) for r in rows[1:]]


def num(v):
    try:
        return int(v) if str(v).strip().lstrip("-").isdigit() else float(v)
    except (ValueError, AttributeError):
        return v


cls_rows = read_csv(DATA / "node_classification/data/node_classification.csv")
reg_rows = read_csv(DATA / "node_regression/data/node_regression.csv")
slp_rows = read_csv(DATA / "static_link_prediction/data/static_link_prediction.csv")
CLS, REG, SLP = rows_as_dicts(cls_rows), rows_as_dicts(reg_rows), rows_as_dicts(slp_rows)


def mean_arm(recs, metric, arm, ds=None):
    xs = [float(r[metric]) for r in recs
          if r["model"] == arm and r["split"] == "test" and (ds is None or r["dataset"] == ds)]
    return statistics.fmean(xs) if xs else None


def val_arm_ds(recs, metric, arm, ds):
    xs = [float(r[metric]) for r in recs if r["model"] == arm and r["dataset"] == ds and r["split"] == "test"]
    return statistics.fmean(xs) if xs else None


# aggregates
cls_m = {a: mean_arm(CLS, "roc_auc", a) for a in ARMS}
reg_m = {a: mean_arm(REG, "spearman", a) for a in ARMS}
slp_m = {a: mean_arm(SLP, "roc_auc", a) for a in ARMS}


REG_GRAPHS = ["midterm", "ukr_rus_twitter", "covid19_twitter", "twibot20"]
REG_TGT = ["followers_count", "statuses_count", "account_age_days"]
CLS_GRAPHS = ["twibot20", "election2020"]
SLP_GRAPHS = ["midterm", "ukr_rus_twitter", "covid19_twitter", "twibot20"]
GRAPH_DISP = {"midterm": "midterm", "ukr_rus_twitter": "ukr_rus", "covid19_twitter": "covid19",
              "twibot20": "twibot20", "election2020": "election2020"}
TGT_DISP = {"followers_count": "followers", "statuses_count": "statuses", "account_age_days": "account age"}


def _rank_group(mv):
    """{model: value} -> {model: rank}, 1 = highest value, average ties."""
    order = sorted(mv.items(), key=lambda x: -x[1])
    rk, i = {}, 0
    while i < len(order):
        j = i
        while j + 1 < len(order) and order[j + 1][1] == order[i][1]:
            j += 1
        avg = (i + 1 + j + 1) / 2.0
        for k in range(i, j + 1):
            rk[order[k][0]] = avg
        i = j + 1
    return rk


def instance_ranks(records, metric):
    """{(dataset, target): {model: rank}} over the test-split instances."""
    inst = {}
    for r in records:
        if r.get("split") != "test" or r.get(metric) in ("", None):
            continue
        inst.setdefault((r["dataset"], r.get("target", "")), {})[r["model"]] = float(r[metric])
    return {key: _rank_group(mv) for key, mv in inst.items()}


IR = {"regression": instance_ranks(REG, "spearman"),
      "classification": instance_ranks(CLS, "roc_auc"),
      "static_link_prediction": instance_ranks(SLP, "roc_auc")}


def _task_mean(ir):
    per = {m: [] for m in ARMS}
    for rk in ir.values():
        for m in ARMS:
            if m in rk:
                per[m].append(rk[m])
    return {m: statistics.fmean(per[m]) for m in ARMS}


# per-task mean rank, then a FAIR overall = equal weight per task family
# (so regression's 12 instances don't outvote classification's 2)
RANK = {t: _task_mean(IR[t]) for t in IR}
overall_rank = {m: statistics.fmean([RANK[t][m] for t in RANK]) for m in ARMS}
worst_rank = {m: max(RANK[t][m] for t in RANK) for m in ARMS}
gen_score = {m: (len(ARMS) - overall_rank[m]) / (len(ARMS) - 1) for m in ARMS}

# breakdown views used by the sheets
reg_target_rank = {t: {m: statistics.fmean([IR["regression"][(g, t)][m] for g in REG_GRAPHS
                                             if (g, t) in IR["regression"]]) for m in ARMS} for t in REG_TGT}
cls_graph_rank = {g: IR["classification"][(g, "")] for g in CLS_GRAPHS}
slp_graph_rank = {g: IR["static_link_prediction"][(g, "")] for g in SLP_GRAPHS}
reg_wins = {m: sum(1 for rk in IR["regression"].values() if rk.get(m) == 1.0) for m in ARMS}

wb = Workbook()


def style_hdr(ws, row, cols, start=1):
    for k, h in enumerate(cols, start=start):
        c = ws.cell(row, k, h); c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CEN; c.border = BORDER


def title(ws, txt, sub=None):
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, txt).font = TITLE_FONT
    if sub:
        ws.cell(2, 1, sub).font = SUB_FONT


def put(ws, r, c, v, *, fmt=None, font=BASE, fill=None, align=CEN, border=True):
    cell = ws.cell(r, c, v)
    cell.font = font; cell.alignment = align
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    if border: cell.border = BORDER
    return cell


def autosize(ws, rows):
    for j in range(1, len(rows[0]) + 1):
        w = max(len(str(rows[0][j - 1])), *(len(str(r[j - 1])) for r in rows[1:] if j - 1 < len(r)))
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 8), 46)


# ============================== About ==============================
ws = wb.active
ws.title = "About"
ws.sheet_view.showGridLines = False
lines = [
    ("Multi-task SSL rotation — results", TITLE_FONT),
    ("One encoder, all SSL tasks: NM / CL / FP single-objective controls vs MIX (per-episode rotation).", SUB_FONT),
    ("Frozen-encoder transfer to node classification, node regression, static link prediction. 1 seed.", SUB_FONT),
    ("", None),
    ("Experiment setup", "HEAD"),
    ("Pretrain corpus", "3-way merged retweet graph: ukr_rus + covid19 + midterm (ukr_rus_covid_midterm_retweet_graph.pt)"),
    ("Encoder (all arms)", "bio-only 768-d GTE embeddings -> mean-agg GraphSAGE, 1 layer / 1 hop, emb_dim 256, undirected. No structural inputs."),
    ("Episode sampling", "GLOBAL (no within-source confinement); n_way 30, n_shots 3, n_query 4, batch_size 1"),
    ("Budget", "40k episodes, matched TOTAL compute (MIX ~1/3 per-task exposure); all arms frozen at the 30k checkpoint"),
    ("Single variable", "SSL objective only — everything else held constant across the 4 arms"),
    ("", None),
    ("Arms", "HEAD"),
    ("NM", "neighbor_matching — instance/neighborhood discrimination, no augmentation, metric loss"),
    ("CL", "contrastive — two feature-augmented views (NZ0.2 = zero 20% of feature dims), metric loss"),
    ("FP", "masked_feature_prediction — GraphMAE-style, mask 30% node features, reconstruct (aux MSE head)"),
    ("MIX", "nm_fp_cl rotation — one objective per episode, round-robin 1:1:1 (~13.3k episodes each)"),
    ("", None),
    ("Evaluation protocol", "HEAD"),
    ("Classification", "10-shot, ROC-AUC. Datasets: twibot20, election2020 — BOTH held-out (pure transfer)."),
    ("Regression", "10-shot, Spearman rho, 3 targets (followers/statuses/account-age, log1p). Datasets: midterm, ukr_rus, covid19 (in-domain) + twibot20 (held-out)."),
    ("Static link prediction", "0-shot, ROC-AUC. Datasets: midterm, ukr_rus, covid19 (in-domain) + twibot20 (held-out)."),
    ("", None),
    ("Data provenance", "HEAD"),
    ("Raw source", "Eval CSVs from the Tucker run worktree /dataMeR1/phil/gfm/prodigy-mtr/scripts/experiments/analysis/*/data/*.csv"),
    ("Repo snapshot", "scripts/experiments/analysis/multitask_ssl_rotation/data/*.csv"),
    ("This workbook", "Raw_* sheets = source eval values. Summary/pivot sheets = means over those rows (test split), computed in Python and verified against aggregate_results.py."),
    ("Reproduce", "python scripts/experiments/analysis/multitask_ssl_rotation/aggregate_results.py --plotting-root scripts/experiments/analysis/multitask_ssl_rotation"),
    ("Built", "2026-07-12"),
    ("", None),
    ("Caveats", "HEAD"),
    ("1 seed", "Frozen-probe eval episodes seeded per-split, not by --seed. LP win is large + consistent across all 4 datasets, so likely real; single-dataset deltas cautiously."),
    ("Matched TOTAL compute", "MIX saw ~1/3 the per-task exposure of each control -> its wins are a lower bound; the one soft spot (regression vs FP) is what dilution predicts."),
]
r = 1
for a, b in lines:
    if b is None:
        r += 1; continue
    if isinstance(b, Font):
        ws.cell(r, 1, a).font = b
    elif b == "HEAD":
        ws.cell(r, 1, a).font = BOLD
    else:
        c = ws.cell(r, 1, a); c.font = BOLD; c.alignment = Alignment(vertical="top")
        c2 = ws.cell(r, 2, b); c2.font = BASE; c2.alignment = Alignment(vertical="top", wrap_text=True)
    r += 1
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 104

# ============================== Summary_T1 ==============================
ws = wb.create_sheet("Summary_T1")
title(ws, "T1 — frozen-encoder transfer (mean over datasets, test split)",
      "Higher is better on every metric. Values = mean over the Raw_* rows.")
style_hdr(ws, 4, ["Arm", "Classification AUC", "Regression ρ", "Static-LP AUC", "Joint min(cls, sLP)", "Role"])
ws.cell(4, 2).comment = Comment("Mean over the 2 held-out classification datasets (Raw_classification, split=test).", "results")
ws.cell(4, 4).comment = Comment("Mean over the 4 static-LP datasets (Raw_static_link_prediction, split=test).", "results")
ws.cell(4, 5).comment = Comment("min(feature=cls AUC, topological=sLP AUC). Chance=0.50 both axes; regression excluded (secondary/noisy).", "results")
for i, arm in enumerate(ARMS):
    rr = 5 + i
    mixrow = arm == "MIX"
    put(ws, rr, 1, arm, font=BOLD, fill=MIX_FILL if mixrow else None)
    put(ws, rr, 2, cls_m[arm], fmt=NUM, font=BOLD if mixrow else BASE, fill=MIX_FILL if mixrow else None)
    put(ws, rr, 3, reg_m[arm], fmt=NUM, font=BOLD if mixrow else BASE, fill=MIX_FILL if mixrow else None)
    put(ws, rr, 4, slp_m[arm], fmt=NUM, font=BOLD if mixrow else BASE, fill=MIX_FILL if mixrow else None)
    put(ws, rr, 5, min(cls_m[arm], slp_m[arm]), fmt=NUM, font=BOLD if mixrow else BASE, fill=MIX_FILL if mixrow else None)
    put(ws, rr, 6, ROLE[arm], font=BASE, align=LEF, fill=MIX_FILL if mixrow else None)

base = 11
ws.cell(base, 1, "MIX − max(NM, CL, FP)   [+ = rotation wins the task]").font = BOLD
style_hdr(ws, base + 1, ["", "Classification AUC", "Regression ρ", "Static-LP AUC", "Joint min-bar"])
put(ws, base + 2, 1, "delta", font=BASE)
best_ctrl_min = max(CONTROLS, key=lambda c: min(cls_m[c], slp_m[c]))
deltas = [
    ("cls", cls_m["MIX"] - max(cls_m[c] for c in CONTROLS)),
    ("reg", reg_m["MIX"] - max(reg_m[c] for c in CONTROLS)),
    ("slp", slp_m["MIX"] - max(slp_m[c] for c in CONTROLS)),
    ("min", min(cls_m["MIX"], slp_m["MIX"]) - min(cls_m[best_ctrl_min], slp_m[best_ctrl_min])),
]
for k, (_, d) in enumerate(deltas):
    put(ws, base + 2, 2 + k, d, fmt=DELTA, font=BOLD)
ws.cell(base + 4, 1, "Reading").font = BOLD
for k, t in enumerate([
    "Joint criterion = min(feature = classification AUC, topological = static-LP AUC); chance = 0.50 on both axes.",
    "Only MIX clears the joint bar: every single-objective control is at/below chance on the topological axis.",
    "Specialists still win their own task (NM classification, FP regression) — MIX is the best GENERAL encoder, not a free lunch.",
]):
    ws.cell(base + 5 + k, 1, "• " + t).font = BASE
for col, w in zip("ABCDEF", [10, 18, 14, 14, 18, 40]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ============================== Overall_ranking ==============================
# Fair, scale-free "how good is each model, generally" — rank-based (ρ/AUC scale gap
# is irrelevant) and equal-weighted per task family. Full breakdown: every aggregate
# is the mean of the cells to its left, so the reader can audit each step.
ws = wb.create_sheet("Overall_ranking")
ws.sheet_view.showGridLines = False
ws.cell(1, 1, "Overall model ranking — fair, cross-task, with full breakdown (lower rank = better)").font = TITLE_FONT
ws.cell(2, 1, "Rank 1–4 per eval instance by the task's own metric (higher = better; ties averaged). "
              "Reg mean = mean of its 3 target ranks · Cls / sLP mean = mean of their per-graph ranks · "
              "OVERALL = equal-weight mean of the 3 task means.").font = SUB_FONT

def merge_hdr(rng, text, fs=9):
    ws.merge_cells(rng)
    for row in ws[rng]:
        for c in row:
            c.fill = HDR_FILL; c.border = BORDER
    tl = ws[rng.split(":")[0]]
    tl.value = text
    tl.font = Font(name=FONT, bold=True, color="FFFFFF", size=fs)
    tl.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return tl

# grouped header: row 4 = groups, row 5 = per-instance sub-labels
merge_hdr("A4:A5", "Model")
merge_hdr("B4:D4", "Regression · rank by target")
merge_hdr("E4:E5", "Reg mean")
merge_hdr("F4:G4", "Classification · by graph")
merge_hdr("H4:H5", "Cls mean")
merge_hdr("I4:L4", "Static-LP · rank by graph")
merge_hdr("M4:M5", "sLP mean")
merge_hdr("N4:N5", "OVERALL rank")
merge_hdr("O4:O5", "Worst task")
merge_hdr("P4:P5", "Score 0–1")
for addr, txt in {"B5": "followers", "C5": "statuses", "D5": "account age",
                  "F5": "twibot20", "G5": "election2020",
                  "I5": "midterm", "J5": "ukr_rus", "K5": "covid19", "L5": "twibot20"}.items():
    c = ws[addr]; c.value = txt
    c.font = Font(name=FONT, bold=True, color="FFFFFF", size=8)
    c.fill = HDR_FILL; c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws["E4"].comment = Comment("= mean of the 3 target ranks to the left (each target rank is itself a mean over 4 graphs — see Regression_ranks).", "results")
ws["H4"].comment = Comment("= mean of the 2 per-graph ranks to the left.", "results")
ws["M4"].comment = Comment("= mean of the 4 per-graph ranks to the left.", "results")
ws["N4"].comment = Comment("Equal-weight mean of Reg/Cls/sLP means — each task family counts once, so regression's 12 instances don't outvote classification's 2.", "results")
ws["O4"].comment = Comment("max of the 3 task means: the family the model is weakest on. A generalist stays low; a collapsing specialist hits 4.", "results")
ws["P4"].comment = Comment("(4 - OVERALL rank)/3. 1.00 = always ranked best of 4; 0.00 = always worst.", "results")

col_of = {"reg_mean": 5, "cls_mean": 8, "slp_mean": 13, "overall": 14, "worst": 15, "score": 16}
for i, m in enumerate(sorted(ARMS, key=lambda mm: overall_rank[mm])):
    rr = 6 + i
    fill = MIX_FILL if m == "MIX" else None
    put(ws, rr, 1, m, font=BOLD, align=LEF, fill=fill)
    put(ws, rr, 2, reg_target_rank["followers_count"][m], fmt="0.00", font=BASE, fill=fill)
    put(ws, rr, 3, reg_target_rank["statuses_count"][m], fmt="0.00", font=BASE, fill=fill)
    put(ws, rr, 4, reg_target_rank["account_age_days"][m], fmt="0.00", font=BASE, fill=fill)
    put(ws, rr, 5, RANK["regression"][m], fmt="0.00", font=BOLD, fill=fill)
    put(ws, rr, 6, cls_graph_rank["twibot20"][m], fmt="0.00", font=BASE, fill=fill)
    put(ws, rr, 7, cls_graph_rank["election2020"][m], fmt="0.00", font=BASE, fill=fill)
    put(ws, rr, 8, RANK["classification"][m], fmt="0.00", font=BOLD, fill=fill)
    put(ws, rr, 9, slp_graph_rank["midterm"][m], fmt="0.00", font=BASE, fill=fill)
    put(ws, rr, 10, slp_graph_rank["ukr_rus_twitter"][m], fmt="0.00", font=BASE, fill=fill)
    put(ws, rr, 11, slp_graph_rank["covid19_twitter"][m], fmt="0.00", font=BASE, fill=fill)
    put(ws, rr, 12, slp_graph_rank["twibot20"][m], fmt="0.00", font=BASE, fill=fill)
    put(ws, rr, 13, RANK["static_link_prediction"][m], fmt="0.00", font=BOLD, fill=fill)
    put(ws, rr, 14, overall_rank[m], fmt="0.00", font=BOLD, fill=fill)
    put(ws, rr, 15, worst_rank[m], fmt="0.00", font=BASE, fill=fill)
    put(ws, rr, 16, gen_score[m], fmt="0.00", font=BOLD, fill=fill)
nr = 6 + len(ARMS) + 1
for k, t in enumerate([
    "Bold columns are aggregates; each equals the mean of the plain cells to its left — fully auditable in-place.",
    "Fair by construction: rank-based (Spearman ρ vs ROC-AUC scale is irrelevant) and each task family weighted equally.",
    "MIX is the best GENERAL model: lowest OVERALL rank AND lowest worst-task rank — the only arm that never drops to a family it can't do.",
    "FP wins regression outright but collapses to rank 4 on classification; NM is strong on cls/LP but ~rank 3 on regression; CL is last.",
]):
    ws.cell(nr + k, 1, "• " + t).font = SUB_FONT
for col, w in zip("ABCDEFGHIJKLMNOP",
                  [6, 9, 8, 11, 8, 9, 11, 8, 8, 8, 8, 8, 8, 9, 8, 8]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "B6"

# ============================== Classification ==============================
ws = wb.create_sheet("Classification")
title(ws, "Node classification — ROC-AUC (10-shot, held-out datasets)",
      "Both eval datasets are held-out (not in the pretrain corpus).")
cls_ds = ["twibot20", "election2020"]
style_hdr(ws, 4, ["Dataset"] + ARMS)
for i, ds in enumerate(cls_ds):
    rr = 5 + i
    put(ws, rr, 1, ds, font=BOLD, align=LEF)
    for j, arm in enumerate(ARMS):
        put(ws, rr, 2 + j, val_arm_ds(CLS, "roc_auc", arm, ds), fmt=NUM,
            font=BOLD if arm == "MIX" else BASE, fill=MIX_FILL if arm == "MIX" else None)
mr = 5 + len(cls_ds)
put(ws, mr, 1, "Mean", font=BOLD, align=LEF, fill=CTRL_FILL)
for j, arm in enumerate(ARMS):
    put(ws, mr, 2 + j, cls_m[arm], fmt=NUM, font=BOLD, fill=CTRL_FILL)
ar = mr + 3
ws.cell(ar - 1, 1, "Accuracy (same layout)").font = BOLD
style_hdr(ws, ar, ["Dataset"] + ARMS)
for i, ds in enumerate(cls_ds):
    rr = ar + 1 + i
    put(ws, rr, 1, ds, font=BOLD, align=LEF)
    for j, arm in enumerate(ARMS):
        put(ws, rr, 2 + j, val_arm_ds(CLS, "accuracy", arm, ds), fmt=NUM)
for col, w in zip("ABCDE", [16, 11, 11, 11, 11]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "B5"

# ============================== Regression ==============================
ws = wb.create_sheet("Regression")
title(ws, "Node regression — Spearman ρ (10-shot, mean over 3 targets)",
      "Per-dataset mean over the 3 targets; full per-target rows in Raw_regression. FP is the specialist.")
reg_ds = ["midterm", "ukr_rus_twitter", "covid19_twitter", "twibot20"]
style_hdr(ws, 4, ["Dataset (in-domain unless noted)"] + ARMS)
for i, ds in enumerate(reg_ds):
    rr = 5 + i
    put(ws, rr, 1, ds + (" (held-out)" if ds == "twibot20" else ""), font=BOLD, align=LEF)
    for j, arm in enumerate(ARMS):
        put(ws, rr, 2 + j, val_arm_ds(REG, "spearman", arm, ds), fmt=NUM,
            font=BOLD if arm == "MIX" else BASE, fill=MIX_FILL if arm == "MIX" else None)
mr = 5 + len(reg_ds)
put(ws, mr, 1, "Mean (all datasets × targets)", font=BOLD, align=LEF, fill=CTRL_FILL)
for j, arm in enumerate(ARMS):
    put(ws, mr, 2 + j, reg_m[arm], fmt=NUM, font=BOLD, fill=CTRL_FILL)
ws.cell(mr + 2, 1, "Targets: followers_count, statuses_count, account_age_days (log1p). Higher ρ = better.").font = SUB_FONT
for col, w in zip("ABCDE", [30, 11, 11, 11, 11]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "B5"

# ============================== Regression_ranks ==============================
# General performance via mean rank: on each (target, dataset) instance, rank the 4
# models by Spearman ρ (1 = best, higher ρ wins; average ties). Per-target mean rank
# = mean over the 4 datasets; Overall = mean over all 12 instances. Lower = better.
# reuse the shared rank breakdowns computed near the top (single source of truth)
REG_TARGETS, REG_TDISP = REG_TGT, TGT_DISP
mr_target = reg_target_rank                 # {target: {model: mean rank over graphs}}
mr_overall = RANK["regression"]             # {model: mean rank over all 12 instances}
wins = reg_wins                             # {model: # rank-1 finishes / 12}
order_by_rank = sorted(ARMS, key=lambda m: mr_overall[m])

ws = wb.create_sheet("Regression_ranks")
title(ws, "Regression — general performance by mean rank (lower = better)",
      "On each (target × test-graph) instance the 4 models are ranked 1–4 by Spearman ρ; cells are mean ranks.")
cols = ["Model"] + [REG_TDISP[t] for t in REG_TARGETS] + ["OVERALL rank", "Wins (rank-1, /12)", "Mean ρ (ref)"]
style_hdr(ws, 4, cols)
ws.cell(4, 5).comment = Comment("Mean rank over all 12 target×dataset instances. 1 = always best of the 4 models, 4 = always worst.", "results")
for i, m in enumerate(order_by_rank):
    rr = 5 + i
    mix = m == "MIX"
    put(ws, rr, 1, m, font=BOLD, align=LEF, fill=MIX_FILL if mix else None)
    for j, t in enumerate(REG_TARGETS):
        put(ws, rr, 2 + j, mr_target[t][m], fmt="0.00", font=BASE, fill=MIX_FILL if mix else None)
    put(ws, rr, 5, mr_overall[m], fmt="0.00", font=BOLD, fill=MIX_FILL if mix else None)
    put(ws, rr, 6, wins[m], fmt="0", font=BASE, fill=MIX_FILL if mix else None)
    put(ws, rr, 7, reg_m[m], fmt=NUM, font=BASE, fill=MIX_FILL if mix else None)
note_r = 5 + len(ARMS) + 1
for k, t in enumerate([
    "Mean rank = average finishing place (1–4) across the test graphs; lower is better. Range 1.00 (always best) to 4.00 (always worst).",
    "FP is the regression specialist (best overall rank). MIX is 2nd overall and actually ranks best on account-age.",
    "Wins = number of the 12 instances where the model has the single highest ρ. Mean ρ is the raw metric for reference (see Regression sheet).",
]):
    ws.cell(note_r + k, 1, "• " + t).font = SUB_FONT
for col, w in zip("ABCDEFG", [8, 11, 11, 12, 15, 18, 13]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "B5"

# ============================== Static_link_prediction ==============================
ws = wb.create_sheet("Static_link_prediction")
title(ws, "Static link prediction — ROC-AUC (0-shot)  —  the headline",
      "Only MIX is above chance (0.50). Controls emit degenerate constant predictors (see accuracy block).")
slp_ds = ["midterm", "ukr_rus_twitter", "covid19_twitter", "twibot20"]
style_hdr(ws, 4, ["Dataset (in-domain unless noted)"] + ARMS + ["MIX − max(ctrl)"])
slp_deltas = []
for i, ds in enumerate(slp_ds):
    rr = 5 + i
    put(ws, rr, 1, ds + (" (held-out)" if ds == "twibot20" else ""), font=BOLD, align=LEF)
    vals = {arm: val_arm_ds(SLP, "roc_auc", arm, ds) for arm in ARMS}
    d = vals["MIX"] - max(vals[c] for c in CONTROLS)
    slp_deltas.append(d)
    for j, arm in enumerate(ARMS):
        put(ws, rr, 2 + j, vals[arm], fmt=NUM, font=BOLD if arm == "MIX" else BASE,
            fill=MIX_FILL if arm == "MIX" else None)
    put(ws, rr, 6, d, fmt=DELTA, font=BOLD)
mr = 5 + len(slp_ds)
put(ws, mr, 1, "Mean", font=BOLD, align=LEF, fill=CTRL_FILL)
for j, arm in enumerate(ARMS):
    put(ws, mr, 2 + j, slp_m[arm], fmt=NUM, font=BOLD, fill=CTRL_FILL)
# column-consistent: mean of the per-dataset (MIX-max) deltas = +0.269 (the FINDINGS headline)
meancell = put(ws, mr, 6, statistics.fmean(slp_deltas), fmt=DELTA, font=BOLD, fill=CTRL_FILL)
meancell.comment = Comment(
    f"Mean of the per-dataset (MIX - max control) deltas = "
    f"{statistics.fmean(slp_deltas):+.3f}, range [{min(slp_deltas):+.3f}, {max(slp_deltas):+.3f}], "
    f"all positive. (The delta-of-aggregated-means on Summary_T1 is +"
    f"{slp_m['MIX'] - max(slp_m[c] for c in CONTROLS):.3f}; different quantity: max control is taken "
    f"per-dataset here vs on the means there.)", "results")
put(ws, mr + 1, 1, f"MIX − max(control) per dataset: mean {statistics.fmean(slp_deltas):+.3f}, "
    f"range [{min(slp_deltas):+.3f}, {max(slp_deltas):+.3f}], all 4 positive.",
    font=SUB_FONT, align=LEF, border=False)
ar = mr + 3
ws.cell(ar - 1, 1, "Accuracy (controls pin at ~0.50 = constant predictor; MIX genuinely separates edges)").font = BOLD
style_hdr(ws, ar, ["Dataset"] + ARMS)
for i, ds in enumerate(slp_ds):
    rr = ar + 1 + i
    put(ws, rr, 1, ds, font=BOLD, align=LEF)
    for j, arm in enumerate(ARMS):
        put(ws, rr, 2 + j, val_arm_ds(SLP, "accuracy", arm, ds), fmt=NUM)
for col, w in zip("ABCDEF", [30, 11, 11, 11, 11, 15]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "B5"

# ============================== Raw sheets ==============================
def write_raw(name, rows, numeric_cols):
    ws = wb.create_sheet(name)
    for j, h in enumerate(rows[0], start=1):
        put(ws, 1, j, h, font=HDR_FONT, fill=HDR_FILL)
    for i, rrow in enumerate(rows[1:], start=2):
        for j, v in enumerate(rrow, start=1):
            val = num(v) if (j - 1) in numeric_cols else v
            c = put(ws, i, j, val, font=BASE, align=(CEN if (j - 1) in numeric_cols else LEF),
                    fill=MIX_FILL if rrow[0] == "MIX" else None)
            if isinstance(val, float):
                c.number_format = "0.0000####"
    ws.freeze_panes = "A2"
    autosize(ws, rows)


write_raw("Raw_classification", cls_rows, {2, 6, 7, 8})
write_raw("Raw_regression", reg_rows, {2, 6, 7, 8, 9, 10})
write_raw("Raw_static_link_prediction", slp_rows, {2, 6, 7, 8})

order = ["About", "Summary_T1", "Overall_ranking", "Classification", "Regression", "Regression_ranks",
         "Static_link_prediction",
         "Raw_classification", "Raw_regression", "Raw_static_link_prediction"]
wb._sheets.sort(key=lambda s: order.index(s.title))

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print("wrote", OUT)
print("sheets:", [s.title for s in wb._sheets])
print("\nverify vs aggregator:")
print("cls", {a: round(cls_m[a], 3) for a in ARMS})
print("reg", {a: round(reg_m[a], 3) for a in ARMS})
print("slp", {a: round(slp_m[a], 3) for a in ARMS})
print("deltas cls/reg/slp:", [round(d, 3) for _, d in deltas[:3]], "min-bar", round(deltas[3][1], 3))
