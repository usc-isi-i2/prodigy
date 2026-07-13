#!/usr/bin/env python3
"""Build the multitask_ssl_rotation results workbook.

Raw eval rows -> Raw_* sheets (the source values). Summary/pivot sheets hold values
computed in Python from those same raw rows (test split), verified against
aggregate_results.py. Values (not live formulas) because LibreOffice is unavailable
here to validate cross-sheet formulas; the raw sheets keep everything auditable and
each derived sheet documents its provenance.
"""
import csv
import statistics
from pathlib import Path
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
REPO = HERE.parents[2]  # scripts/experiments/multitask_ssl_rotation -> repo root
DATA = REPO / "scripts/plotting/multitask_ssl_rotation"
OUT = HERE / "multitask_ssl_rotation_results.xlsx"

ARMS = ["NM", "CL", "FP", "MIX"]
CONTROLS = ["NM", "CL", "FP"]
ROLE = {
    "NM": "control — classification specialist",
    "CL": "control — weak everywhere",
    "FP": "control — regression specialist",
    "MIX": "TREATMENT — per-episode rotation",
}

FONT = "Arial"
INK = "202124"
HDR_FILL = PatternFill("solid", fgColor="34405A")
HDR_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color=INK)
SUB_FONT = Font(name=FONT, italic=True, size=10, color="5F6368")
BOLD = Font(name=FONT, bold=True, color=INK, size=11)
BASE = Font(name=FONT, color=INK, size=11)
MIX_FILL = PatternFill("solid", fgColor="ECEBFA")
CTRL_FILL = PatternFill("solid", fgColor="F5F6F8")
thin = Side(style="thin", color="D0D3D8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
NUM = "0.000"
DELTA = "+0.000;\\-0.000;0.000"
CEN = Alignment(horizontal="center", vertical="center")
LEF = Alignment(horizontal="left", vertical="center")


def read_csv(p):
    with open(p) as fh:
        return list(csv.reader(fh))


def rows_as_dicts(rows):
    h = rows[0]
    return [dict(zip(h, r)) for r in rows[1:]]


def num(v):
    try:
        return int(v) if str(v).strip().lstrip("-").isdigit() else float(v)
    except (ValueError, AttributeError):
        return v


cls_rows = read_csv(DATA / "node_classification/data/node_classification.csv")
reg_rows = read_csv(DATA / "node_regression/data/node_regression.csv")
slp_rows = read_csv(DATA / "static_link_prediction/data/static_link_prediction.csv")
CLS, REG, SLP = rows_as_dicts(cls_rows), rows_as_dicts(reg_rows), rows_as_dicts(slp_rows)


def mean_arm(recs, metric, arm, ds=None):
    xs = [float(r[metric]) for r in recs
          if r["model"] == arm and r["split"] == "test" and (ds is None or r["dataset"] == ds)]
    return statistics.fmean(xs) if xs else None


def val_arm_ds(recs, metric, arm, ds):
    xs = [float(r[metric]) for r in recs if r["model"] == arm and r["dataset"] == ds and r["split"] == "test"]
    return statistics.fmean(xs) if xs else None


# aggregates
cls_m = {a: mean_arm(CLS, "roc_auc", a) for a in ARMS}
reg_m = {a: mean_arm(REG, "spearman", a) for a in ARMS}
slp_m = {a: mean_arm(SLP, "roc_auc", a) for a in ARMS}

wb = Workbook()


def style_hdr(ws, row, cols, start=1):
    for k, h in enumerate(cols, start=start):
        c = ws.cell(row, k, h); c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CEN; c.border = BORDER


def title(ws, txt, sub=None):
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, txt).font = TITLE_FONT
    if sub:
        ws.cell(2, 1, sub).font = SUB_FONT


def put(ws, r, c, v, *, fmt=None, font=BASE, fill=None, align=CEN, border=True):
    cell = ws.cell(r, c, v)
    cell.font = font; cell.alignment = align
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    if border: cell.border = BORDER
    return cell


def autosize(ws, rows):
    for j in range(1, len(rows[0]) + 1):
        w = max(len(str(rows[0][j - 1])), *(len(str(r[j - 1])) for r in rows[1:] if j - 1 < len(r)))
        ws.column_dimensions[get_column_letter(j)].width = min(max(w + 2, 8), 46)


# ============================== About ==============================
ws = wb.active
ws.title = "About"
ws.sheet_view.showGridLines = False
lines = [
    ("Multi-task SSL rotation — results", TITLE_FONT),
    ("One encoder, all SSL tasks: NM / CL / FP single-objective controls vs MIX (per-episode rotation).", SUB_FONT),
    ("Frozen-encoder transfer to node classification, node regression, static link prediction. 1 seed.", SUB_FONT),
    ("", None),
    ("Experiment setup", "HEAD"),
    ("Pretrain corpus", "3-way merged retweet graph: ukr_rus + covid19 + midterm (ukr_rus_covid_midterm_retweet_graph.pt)"),
    ("Encoder (all arms)", "bio-only 768-d GTE embeddings -> mean-agg GraphSAGE, 1 layer / 1 hop, emb_dim 256, undirected. No structural inputs."),
    ("Episode sampling", "GLOBAL (no within-source confinement); n_way 30, n_shots 3, n_query 4, batch_size 1"),
    ("Budget", "40k episodes, matched TOTAL compute (MIX ~1/3 per-task exposure); all arms frozen at the 30k checkpoint"),
    ("Single variable", "SSL objective only — everything else held constant across the 4 arms"),
    ("", None),
    ("Arms", "HEAD"),
    ("NM", "neighbor_matching — instance/neighborhood discrimination, no augmentation, metric loss"),
    ("CL", "contrastive — two feature-augmented views (NZ0.2 = zero 20% of feature dims), metric loss"),
    ("FP", "masked_feature_prediction — GraphMAE-style, mask 30% node features, reconstruct (aux MSE head)"),
    ("MIX", "nm_fp_cl rotation — one objective per episode, round-robin 1:1:1 (~13.3k episodes each)"),
    ("", None),
    ("Evaluation protocol", "HEAD"),
    ("Classification", "10-shot, ROC-AUC. Datasets: twibot20, election2020 — BOTH held-out (pure transfer)."),
    ("Regression", "10-shot, Spearman rho, 3 targets (followers/statuses/account-age, log1p). Datasets: midterm, ukr_rus, covid19 (in-domain) + twibot20 (held-out)."),
    ("Static link prediction", "0-shot, ROC-AUC. Datasets: midterm, ukr_rus, covid19 (in-domain) + twibot20 (held-out)."),
    ("", None),
    ("Data provenance", "HEAD"),
    ("Raw source", "Eval CSVs from the Tucker run worktree /dataMeR1/phil/gfm/prodigy-mtr/scripts/plotting/*/data/*.csv"),
    ("Repo snapshot", "scripts/plotting/multitask_ssl_rotation/{node_classification,node_regression,static_link_prediction}/data/*.csv"),
    ("This workbook", "Raw_* sheets = source eval values. Summary/pivot sheets = means over those rows (test split), computed in Python and verified against aggregate_results.py."),
    ("Reproduce", "python scripts/experiments/multitask_ssl_rotation/aggregate_results.py --plotting-root scripts/plotting/multitask_ssl_rotation"),
    ("Built", "2026-07-12"),
    ("", None),
    ("Caveats", "HEAD"),
    ("1 seed", "Frozen-probe eval episodes seeded per-split, not by --seed. LP win is large + consistent across all 4 datasets, so likely real; single-dataset deltas cautiously."),
    ("Matched TOTAL compute", "MIX saw ~1/3 the per-task exposure of each control -> its wins are a lower bound; the one soft spot (regression vs FP) is what dilution predicts."),
]
r = 1
for a, b in lines:
    if b is None:
        r += 1; continue
    if isinstance(b, Font):
        ws.cell(r, 1, a).font = b
    elif b == "HEAD":
        ws.cell(r, 1, a).font = BOLD
    else:
        c = ws.cell(r, 1, a); c.font = BOLD; c.alignment = Alignment(vertical="top")
        c2 = ws.cell(r, 2, b); c2.font = BASE; c2.alignment = Alignment(vertical="top", wrap_text=True)
    r += 1
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 104

# ============================== Summary_T1 ==============================
ws = wb.create_sheet("Summary_T1")
title(ws, "T1 — frozen-encoder transfer (mean over datasets, test split)",
      "Higher is better on every metric. Values = mean over the Raw_* rows.")
style_hdr(ws, 4, ["Arm", "Classification AUC", "Regression ρ", "Static-LP AUC", "Joint min(cls, sLP)", "Role"])
ws.cell(4, 2).comment = Comment("Mean over the 2 held-out classification datasets (Raw_classification, split=test).", "results")
ws.cell(4, 4).comment = Comment("Mean over the 4 static-LP datasets (Raw_static_link_prediction, split=test).", "results")
ws.cell(4, 5).comment = Comment("min(feature=cls AUC, topological=sLP AUC). Chance=0.50 both axes; regression excluded (secondary/noisy).", "results")
for i, arm in enumerate(ARMS):
    rr = 5 + i
    mixrow = arm == "MIX"
    put(ws, rr, 1, arm, font=BOLD, fill=MIX_FILL if mixrow else None)
    put(ws, rr, 2, cls_m[arm], fmt=NUM, font=BOLD if mixrow else BASE, fill=MIX_FILL if mixrow else None)
    put(ws, rr, 3, reg_m[arm], fmt=NUM, font=BOLD if mixrow else BASE, fill=MIX_FILL if mixrow else None)
    put(ws, rr, 4, slp_m[arm], fmt=NUM, font=BOLD if mixrow else BASE, fill=MIX_FILL if mixrow else None)
    put(ws, rr, 5, min(cls_m[arm], slp_m[arm]), fmt=NUM, font=BOLD if mixrow else BASE, fill=MIX_FILL if mixrow else None)
    put(ws, rr, 6, ROLE[arm], font=BASE, align=LEF, fill=MIX_FILL if mixrow else None)

base = 11
ws.cell(base, 1, "MIX − max(NM, CL, FP)   [+ = rotation wins the task]").font = BOLD
style_hdr(ws, base + 1, ["", "Classification AUC", "Regression ρ", "Static-LP AUC", "Joint min-bar"])
put(ws, base + 2, 1, "delta", font=BASE)
best_ctrl_min = max(CONTROLS, key=lambda c: min(cls_m[c], slp_m[c]))
deltas = [
    ("cls", cls_m["MIX"] - max(cls_m[c] for c in CONTROLS)),
    ("reg", reg_m["MIX"] - max(reg_m[c] for c in CONTROLS)),
    ("slp", slp_m["MIX"] - max(slp_m[c] for c in CONTROLS)),
    ("min", min(cls_m["MIX"], slp_m["MIX"]) - min(cls_m[best_ctrl_min], slp_m[best_ctrl_min])),
]
for k, (_, d) in enumerate(deltas):
    put(ws, base + 2, 2 + k, d, fmt=DELTA, font=BOLD)
ws.cell(base + 4, 1, "Reading").font = BOLD
for k, t in enumerate([
    "Joint criterion = min(feature = classification AUC, topological = static-LP AUC); chance = 0.50 on both axes.",
    "Only MIX clears the joint bar: every single-objective control is at/below chance on the topological axis.",
    "Specialists still win their own task (NM classification, FP regression) — MIX is the best GENERAL encoder, not a free lunch.",
]):
    ws.cell(base + 5 + k, 1, "• " + t).font = BASE
for col, w in zip("ABCDEF", [10, 18, 14, 14, 18, 40]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

# ============================== Classification ==============================
ws = wb.create_sheet("Classification")
title(ws, "Node classification — ROC-AUC (10-shot, held-out datasets)",
      "Both eval datasets are held-out (not in the pretrain corpus).")
cls_ds = ["twibot20", "election2020"]
style_hdr(ws, 4, ["Dataset"] + ARMS)
for i, ds in enumerate(cls_ds):
    rr = 5 + i
    put(ws, rr, 1, ds, font=BOLD, align=LEF)
    for j, arm in enumerate(ARMS):
        put(ws, rr, 2 + j, val_arm_ds(CLS, "roc_auc", arm, ds), fmt=NUM,
            font=BOLD if arm == "MIX" else BASE, fill=MIX_FILL if arm == "MIX" else None)
mr = 5 + len(cls_ds)
put(ws, mr, 1, "Mean", font=BOLD, align=LEF, fill=CTRL_FILL)
for j, arm in enumerate(ARMS):
    put(ws, mr, 2 + j, cls_m[arm], fmt=NUM, font=BOLD, fill=CTRL_FILL)
ar = mr + 3
ws.cell(ar - 1, 1, "Accuracy (same layout)").font = BOLD
style_hdr(ws, ar, ["Dataset"] + ARMS)
for i, ds in enumerate(cls_ds):
    rr = ar + 1 + i
    put(ws, rr, 1, ds, font=BOLD, align=LEF)
    for j, arm in enumerate(ARMS):
        put(ws, rr, 2 + j, val_arm_ds(CLS, "accuracy", arm, ds), fmt=NUM)
for col, w in zip("ABCDE", [16, 11, 11, 11, 11]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "B5"

# ============================== Regression ==============================
ws = wb.create_sheet("Regression")
title(ws, "Node regression — Spearman ρ (10-shot, mean over 3 targets)",
      "Per-dataset mean over the 3 targets; full per-target rows in Raw_regression. FP is the specialist.")
reg_ds = ["midterm", "ukr_rus_twitter", "covid19_twitter", "twibot20"]
style_hdr(ws, 4, ["Dataset (in-domain unless noted)"] + ARMS)
for i, ds in enumerate(reg_ds):
    rr = 5 + i
    put(ws, rr, 1, ds + (" (held-out)" if ds == "twibot20" else ""), font=BOLD, align=LEF)
    for j, arm in enumerate(ARMS):
        put(ws, rr, 2 + j, val_arm_ds(REG, "spearman", arm, ds), fmt=NUM,
            font=BOLD if arm == "MIX" else BASE, fill=MIX_FILL if arm == "MIX" else None)
mr = 5 + len(reg_ds)
put(ws, mr, 1, "Mean (all datasets × targets)", font=BOLD, align=LEF, fill=CTRL_FILL)
for j, arm in enumerate(ARMS):
    put(ws, mr, 2 + j, reg_m[arm], fmt=NUM, font=BOLD, fill=CTRL_FILL)
ws.cell(mr + 2, 1, "Targets: followers_count, statuses_count, account_age_days (log1p). Higher ρ = better.").font = SUB_FONT
for col, w in zip("ABCDE", [30, 11, 11, 11, 11]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "B5"

# ============================== Static_link_prediction ==============================
ws = wb.create_sheet("Static_link_prediction")
title(ws, "Static link prediction — ROC-AUC (0-shot)  —  the headline",
      "Only MIX is above chance (0.50). Controls emit degenerate constant predictors (see accuracy block).")
slp_ds = ["midterm", "ukr_rus_twitter", "covid19_twitter", "twibot20"]
style_hdr(ws, 4, ["Dataset (in-domain unless noted)"] + ARMS + ["MIX − max(ctrl)"])
slp_deltas = []
for i, ds in enumerate(slp_ds):
    rr = 5 + i
    put(ws, rr, 1, ds + (" (held-out)" if ds == "twibot20" else ""), font=BOLD, align=LEF)
    vals = {arm: val_arm_ds(SLP, "roc_auc", arm, ds) for arm in ARMS}
    d = vals["MIX"] - max(vals[c] for c in CONTROLS)
    slp_deltas.append(d)
    for j, arm in enumerate(ARMS):
        put(ws, rr, 2 + j, vals[arm], fmt=NUM, font=BOLD if arm == "MIX" else BASE,
            fill=MIX_FILL if arm == "MIX" else None)
    put(ws, rr, 6, d, fmt=DELTA, font=BOLD)
mr = 5 + len(slp_ds)
put(ws, mr, 1, "Mean", font=BOLD, align=LEF, fill=CTRL_FILL)
for j, arm in enumerate(ARMS):
    put(ws, mr, 2 + j, slp_m[arm], fmt=NUM, font=BOLD, fill=CTRL_FILL)
# column-consistent: mean of the per-dataset (MIX-max) deltas = +0.269 (the FINDINGS headline)
meancell = put(ws, mr, 6, statistics.fmean(slp_deltas), fmt=DELTA, font=BOLD, fill=CTRL_FILL)
meancell.comment = Comment(
    f"Mean of the per-dataset (MIX - max control) deltas = "
    f"{statistics.fmean(slp_deltas):+.3f}, range [{min(slp_deltas):+.3f}, {max(slp_deltas):+.3f}], "
    f"all positive. (The delta-of-aggregated-means on Summary_T1 is +"
    f"{slp_m['MIX'] - max(slp_m[c] for c in CONTROLS):.3f}; different quantity: max control is taken "
    f"per-dataset here vs on the means there.)", "results")
put(ws, mr + 1, 1, f"MIX − max(control) per dataset: mean {statistics.fmean(slp_deltas):+.3f}, "
    f"range [{min(slp_deltas):+.3f}, {max(slp_deltas):+.3f}], all 4 positive.",
    font=SUB_FONT, align=LEF, border=False)
ar = mr + 3
ws.cell(ar - 1, 1, "Accuracy (controls pin at ~0.50 = constant predictor; MIX genuinely separates edges)").font = BOLD
style_hdr(ws, ar, ["Dataset"] + ARMS)
for i, ds in enumerate(slp_ds):
    rr = ar + 1 + i
    put(ws, rr, 1, ds, font=BOLD, align=LEF)
    for j, arm in enumerate(ARMS):
        put(ws, rr, 2 + j, val_arm_ds(SLP, "accuracy", arm, ds), fmt=NUM)
for col, w in zip("ABCDEF", [30, 11, 11, 11, 11, 15]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "B5"

# ============================== Raw sheets ==============================
def write_raw(name, rows, numeric_cols):
    ws = wb.create_sheet(name)
    for j, h in enumerate(rows[0], start=1):
        put(ws, 1, j, h, font=HDR_FONT, fill=HDR_FILL)
    for i, rrow in enumerate(rows[1:], start=2):
        for j, v in enumerate(rrow, start=1):
            val = num(v) if (j - 1) in numeric_cols else v
            c = put(ws, i, j, val, font=BASE, align=(CEN if (j - 1) in numeric_cols else LEF),
                    fill=MIX_FILL if rrow[0] == "MIX" else None)
            if isinstance(val, float):
                c.number_format = "0.0000####"
    ws.freeze_panes = "A2"
    autosize(ws, rows)


write_raw("Raw_classification", cls_rows, {2, 6, 7, 8})
write_raw("Raw_regression", reg_rows, {2, 6, 7, 8, 9, 10})
write_raw("Raw_static_link_prediction", slp_rows, {2, 6, 7, 8})

order = ["About", "Summary_T1", "Classification", "Regression", "Static_link_prediction",
         "Raw_classification", "Raw_regression", "Raw_static_link_prediction"]
wb._sheets.sort(key=lambda s: order.index(s.title))

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print("wrote", OUT)
print("sheets:", [s.title for s in wb._sheets])
print("\nverify vs aggregator:")
print("cls", {a: round(cls_m[a], 3) for a in ARMS})
print("reg", {a: round(reg_m[a], 3) for a in ARMS})
print("slp", {a: round(slp_m[a], 3) for a in ARMS})
print("deltas cls/reg/slp:", [round(d, 3) for _, d in deltas[:3]], "min-bar", round(deltas[3][1], 3))
